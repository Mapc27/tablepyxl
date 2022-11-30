from lxml import html
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from premailer import Premailer

from tablepyxl.style import Table, get_side


def string_to_int(s):
    if s.isdigit():
        return int(s)
    return 0


def get_tables(doc):
    tree = html.fromstring(doc)
    comments = tree.xpath('//comment()')
    for comment in comments:
        comment.drop_tag()
    tables = tree.xpath('//table[not(ancestor::table)]')
    result = [Table(table) for table in tables]
    return result


class TableToWorksheet:
    def __init__(self, worksheet, table):
        self.worksheet = worksheet
        self.table = table

    def write_cell(self, table_cell, row, column):
        cell = self.worksheet.cell(row=row, column=column)

        colspan = string_to_int(table_cell.element.get("colspan", "1"))
        rowspan = string_to_int(table_cell.element.get("rowspan", "1"))
        try:
            table_cell.value[-1].text = table_cell.value[-1].text.replace('\n', '')
        except IndexError:
            pass
        cell_arr = [len(i) for i in str(table_cell.value).split('\n')]
        height_cell = int(len(cell_arr) * 15)
        width_cell = max(cell_arr) + 2

        while isinstance(cell, MergedCell):
            width = max(
                self.worksheet.column_dimensions[get_column_letter(column)].width or 0,
                width_cell // colspan + 1
            )
            self.worksheet.column_dimensions[get_column_letter(column)].width = width

            table_cell.format(cell)
            column += 1
            cell = self.worksheet.cell(row=row, column=column)

        if rowspan > 1 or colspan > 1:
            self.worksheet.merge_cells(start_row=row, start_column=column,
                                       end_row=row + rowspan - 1, end_column=column + colspan - 1)

        cell.value = table_cell.value
        table_cell.format(cell)

        width = max(self.worksheet.column_dimensions[get_column_letter(column)].width, width_cell // colspan + 1)
        self.worksheet.column_dimensions[get_column_letter(column)].width = width
        self.worksheet.row_dimensions[row].height = max(self.worksheet.row_dimensions[row].height or 15, height_cell)

        return column

    def write_rows(self, row, column=1):
        elem = self.table.body

        initial_row = row
        initial_column = column

        for i in range(1, len(elem.rows) + 1):
            self.worksheet.row_dimensions[i].height = 15

        for table_row in elem.rows:
            column = initial_column
            for table_cell in table_row.cells:
                column = self.write_cell(table_cell, row, column)
                column += 1
            row += 1

        self.set_external_top_border(start_column=initial_column, end_column=column - 1, row=initial_row)
        self.set_external_bottom_border(start_column=initial_column, end_column=column - 1, row=row - 1)
        self.set_external_left_border(start_row=initial_row, end_row=row - 1, column=initial_column)
        self.set_external_right_border(start_row=initial_row, end_row=row - 1, column=column - 1)

        return row

    def set_external_top_border(self, start_column, end_column, row):
        top = get_side(self.table.style_dict, 'top')
        if top['border_style'] or top['color']:
            top = Side(**top)
            for column in range(start_column, end_column + 1):
                cell = self.worksheet.cell(row, column)
                bottom = Side(**cell.border.bottom.__dict__)
                left = Side(**cell.border.left.__dict__)
                right = Side(**cell.border.right.__dict__)

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def set_external_bottom_border(self, start_column, end_column, row):
        bottom = get_side(self.table.style_dict, 'bottom')
        if bottom['border_style'] or bottom['color']:
            bottom = Side(**bottom)
            for column in range(start_column, end_column + 1):
                cell = self.worksheet.cell(row, column)
                top = Side(**cell.border.top.__dict__)
                left = Side(**cell.border.left.__dict__)
                right = Side(**cell.border.right.__dict__)

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def set_external_left_border(self, start_row, end_row, column):
        left = get_side(self.table.style_dict, 'left')
        if left['border_style'] or left['color']:
            left = Side(**left)
            for row in range(start_row, end_row + 1):
                cell = self.worksheet.cell(row, column)
                top = Side(**cell.border.top.__dict__)
                bottom = Side(**cell.border.bottom.__dict__)
                right = Side(**cell.border.right.__dict__)

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def set_external_right_border(self, start_row, end_row, column):
        right = get_side(self.table.style_dict, 'right')
        if right['border_style'] or right['color']:
            right = Side(**right)
            for row in range(start_row, end_row + 1):
                cell = self.worksheet.cell(row, column)
                top = Side(**cell.border.top.__dict__)
                left = Side(**cell.border.left.__dict__)
                bottom = Side(**cell.border.bottom.__dict__)

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def tables_to_sheet(tables, wb):
    worksheet = wb.create_sheet()
    row, column = 1, 1
    for table in tables:
        table_to_worksheet = TableToWorksheet(worksheet, table)
        # if table.head:
        #     row = table_to_worksheet.write_rows(worksheet, table.style_dict, row, column)
        if table.body:
            row = table_to_worksheet.write_rows(row, column)
        row += 1


def document_to_workbook(doc, wb=None, base_url=None):
    if not wb:
        wb = Workbook()
        wb.remove(wb.active)

    inline_styles_doc = Premailer(doc, base_url=base_url, remove_classes=False).transform()
    tables = get_tables(inline_styles_doc)
    tables_to_sheet(tables, wb)
    return wb


def document_to_xl(doc, filename, base_url=None):
    doc = doc.replace('<br>', '\n').replace('<br />', '\n')
    wb = document_to_workbook(doc, base_url=base_url)
    wb.save(filename)
