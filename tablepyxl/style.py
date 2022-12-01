import re

from lxml.html import HtmlElement
from openpyxl.cell import cell as openpyxl_cell
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont, Text
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.colors import BLACK
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE

from tablepyxl.exceptions import IntNotFoundException

FORMAT_DATE_MM_DD_YYYY = 'mm/dd/yyyy'

STYLES_CONVERTER_DICT = {
    'bgcolor': 'background-color',
    'align': 'text-align',
    'color': 'color',
    'size': 'font-size'
}

basic_color_names = {
    'aliceblue': '#f0f8ff', 'antiquewhite': '#faebd7', 'aqua': '#00ffff', 'aquamarine': '#7fffd4', 'azure': '#f0ffff',
    'beige': '#f5f5dc', 'bisque': '#ffe4c4', 'black': '#000000', 'blanchedalmond': '#ffebcd', 'blue': '#0000ff',
    'blueviolet': '#8a2be2', 'brown': '#a52a2a', 'burlywood': '#deb887', 'cadetblue': '#5f9ea0',
    'chartreuse': '#7fff00', 'chocolate': '#d2691e', 'coral': '#ff7f50', 'cornflowerblue': '#6495ed',
    'cornsilk': '#fff8dc', 'crimson': '#dc143c', 'cyan': '#00ffff', 'darkblue': '#00008b', 'darkcyan': '#008b8b',
    'darkgoldenrod': '#b8860b', 'darkgray': '#a9a9a9', 'darkgreen': '#006400', 'darkgrey': '#a9a9a9',
    'darkkhaki': '#bdb76b', 'darkmagenta': '#8b008b', 'darkolivegreen': '#556b2f', 'darkorange': '#ff8c00',
    'darkorchid': '#9932cc', 'darkred': '#8b0000', 'darksalmon': '#e9967a', 'darkseagreen': '#8fbc8f',
    'darkslateblue': '#483d8b', 'darkslategray': '#2f4f4f', 'darkslategrey': '#2f4f4f', 'darkturquoise': '#00ced1',
    'darkviolet': '#9400d3', 'deeppink': '#ff1493', 'deepskyblue': '#00bfff', 'dimgray': '#696969',
    'dimgrey': '#696969', 'dodgerblue': '#1e90ff', 'firebrick': '#b22222', 'floralwhite': '#fffaf0',
    'forestgreen': '#228b22', 'fuchsia': '#ff00ff', 'gainsboro': '#dcdcdc', 'ghostwhite': '#f8f8ff',
    'gold': '#ffd700', 'goldenrod': '#daa520', 'gray': '#808080', 'green': '#008000', 'greenyellow': '#adff2f',
    'grey': '#808080', 'honeydew': '#f0fff0', 'hotpink': '#ff69b4', 'indianred': '#cd5c5c', 'indigo': '#4b0082',
    'ivory': '#fffff0', 'khaki': '#f0e68c', 'lavender': '#e6e6fa', 'lavenderblush': '#fff0f5', 'lawngreen': '#7cfc00',
    'lemonchiffon': '#fffacd', 'lightblue': '#add8e6', 'lightcoral': '#f08080', 'lightcyan': '#e0ffff',
    'lightgoldenrodyellow': '#fafad2', 'lightgray': '#d3d3d3', 'lightgreen': '#90ee90', 'lightgrey': '#d3d3d3',
    'lightpink': '#ffb6c1', 'lightsalmon': '#ffa07a', 'lightseagreen': '#20b2aa', 'lightskyblue': '#87cefa',
    'lightslategray': '#778899', 'lightslategrey': '#778899', 'lightsteelblue': '#b0c4de', 'lightyellow': '#ffffe0',
    'lime': '#00ff00', 'limegreen': '#32cd32', 'linen': '#faf0e6', 'magenta': '#ff00ff', 'maroon': '#800000',
    'mediumaquamarine': '#66cdaa', 'mediumblue': '#0000cd', 'mediumorchid': '#ba55d3', 'mediumpurple': '#9370db',
    'mediumseagreen': '#3cb371', 'mediumslateblue': '#7b68ee', 'mediumspringgreen': '#00fa9a',
    'mediumturquoise': '#48d1cc', 'mediumvioletred': '#c71585', 'midnightblue': '#191970', 'mintcream': '#f5fffa',
    'mistyrose': '#ffe4e1', 'moccasin': '#ffe4b5', 'navajowhite': '#ffdead', 'navy': '#000080', 'oldlace': '#fdf5e6',
    'olive': '#808000', 'olivedrab': '#6b8e23', 'orange': '#ffa500', 'orangered': '#ff4500', 'orchid': '#da70d6',
    'palegoldenrod': '#eee8aa', 'palegreen': '#98fb98', 'paleturquoise': '#afeeee', 'palevioletred': '#db7093',
    'papayawhip': '#ffefd5', 'peachpuff': '#ffdab9', 'peru': '#cd853f', 'pink': '#ffc0cb', 'plum': '#dda0dd',
    'powderblue': '#b0e0e6', 'purple': '#800080', 'red': '#ff0000', 'rosybrown': '#bc8f8f', 'royalblue': '#4169e1',
    'saddlebrown': '#8b4513', 'salmon': '#fa8072', 'sandybrown': '#f4a460', 'seagreen': '#2e8b57',
    'seashell': '#fff5ee', 'sienna': '#a0522d', 'silver': '#c0c0c0', 'skyblue': '#87ceeb', 'slateblue': '#6a5acd',
    'slategray': '#708090', 'slategrey': '#708090', 'snow': '#fffafa', 'springgreen': '#00ff7f', 'steelblue': '#4682b4',
    'tan': '#d2b48c', 'teal': '#008080', 'thistle': '#d8bfd8', 'tomato': '#ff6347', 'turquoise': '#40e0d0',
    'violet': '#ee82ee', 'wheat': '#f5deb3', 'white': '#ffffff', 'whitesmoke': '#f5f5f5', 'yellow': '#ffff00',
    'yellowgreen': '#9acd32'
}


def get_hex(color):
    basic_color = basic_color_names.get(color)
    color = basic_color if basic_color is not None else color

    if hasattr(color, 'startswith') and color.startswith('#'):
        color = color[1:]
        if len(color) == 3:
            color = ''.join(2 * c for c in color)
    if not color:
        color = None
    return color


def extract_first_int_from_str(string):
    try:
        return re.findall(r'\d+', string)[0]
    except (IndexError, TypeError):
        raise IntNotFoundException(f"Can't found int value from string = {string}")


def colormap(color):
    """
    Convenience for looking up known colors
    """
    cmap = {'black': BLACK}
    return cmap.get(color, color)


def style_string_to_dict(style):
    def clean_split(string, delim):
        return (s.strip() for s in string.split(delim))

    styles = [clean_split(s, ":") for s in style.split(";") if ":" in s]
    return dict(styles)


def get_side(style_dict, name):
    color = style_dict.get_color('border-{}-color'.format(name))
    style = style_dict.get('border-{}-style'.format(name))
    if style:
        style = style.lower()
    if style in ['dotted', 'dashed', 'double']:
        return {
            'color': color,
            'border_style': style
        }
    if style == 'solid':
        style = 'thin'

        width = style_dict.get('border-{}-width'.format(name))
        width = int(extract_first_int_from_str(width))
        if width == 0:
            style = None
        elif width == 1:
            style = 'thin'
        elif width == 2:
            style = 'medium'
        elif width >= 3:
            style = 'thick'

    return {
        'color': color,
        'border_style': style
    }


known_styles = {}


def get_dimension(dimension):
    if dimension:
        unit = dimension[-2:]
        if unit in ['px', 'em', 'pt', 'in', 'cm']:
            dimension = dimension[:-2]
        dimension = float(dimension)
        if unit == 'em':
            dimension *= 16

    return dimension


def style_dict_to_named_style(style_dict, number_format=None):
    el = style_dict
    parents = ''
    while el.parent is not None:
        parents += str(el.parent)
        el = el.parent

    style_and_format_string = str({
        'style_dict': style_dict,
        'parent': parents,
        'number_format': number_format,
    })
    if style_and_format_string not in known_styles:
        font = Font(
            name=style_dict.get('font-family'),
            size=get_dimension(style_dict.get('font-size')),
            bold=style_dict.get('font-weight') in ['bold', '700'],
            italic=style_dict.get('font-style') == 'italic',
            color=style_dict.get_color('color', None),
        )

        horizontal = style_dict.get('text-align', 'general')
        vertical = style_dict.get('vertical-align', 'top')
        alignment = Alignment(
            horizontal=horizontal if horizontal in (
                'right', 'justify', 'distributed', 'fill', 'centerContinuous', 'center', 'general', 'left') else 'left',
            vertical=vertical if vertical in (
                'bottom', 'center', 'justify', 'top', 'distributed') else 'top',
            # wrap_text=style_dict.get('white-space', 'nowrap') == 'wrap',
            wrap_text=True,
            # indent=get_dimension(style_dict.get('padding')) or 0.0
        )

        bg_color = style_dict.get_color('background-color')
        fg_color = style_dict.get_color('foreground-color', Color())
        fill_type = style_dict.get('fill-type', FILL_SOLID)
        if bg_color and bg_color != 'transparent':
            fill = PatternFill(
                fill_type=fill_type,
                start_color=bg_color,
                end_color=fg_color
            )
        else:
            fill = PatternFill()

        border = Border(
            left=Side(**get_side(style_dict, 'left')),
            right=Side(**get_side(style_dict, 'right')),
            top=Side(**get_side(style_dict, 'top')),
            bottom=Side(**get_side(style_dict, 'bottom')),
            diagonal=Side(**get_side(style_dict, 'diagonal')),
            diagonal_direction=None,
            outline=Side(**get_side(style_dict, 'outline')),
            vertical=None,
            horizontal=None
        )

        name = 'Style {}'.format(len(known_styles) + 1)

        pyxl_style = NamedStyle(
            name=name,
            font=font,
            fill=fill,
            alignment=alignment,
            border=border,
            number_format=number_format
        )

        known_styles[style_and_format_string] = pyxl_style
    return known_styles[style_and_format_string]


class StyleDict(dict):
    def __init__(self, *args, **kwargs):
        self.parent = kwargs.pop('parent', None)
        super(StyleDict, self).__init__(*args, **kwargs)

    def __getitem__(self, item):
        if item in self:
            return super(StyleDict, self).__getitem__(item)
        elif self.parent is not None:
            return self.parent[item]
        else:
            raise KeyError('{} not found'.format(item))

    def __hash__(self):
        return hash(tuple([(k, self.get(k)) for k in self._keys()]))

    def _keys(self):
        yielded = set()
        for k in self.keys():
            yielded.add(k)
            yield k
        if self.parent:
            for k in self.parent._keys():
                if k not in yielded:
                    yielded.add(k)
                    yield k

    def get(self, k, d=None):
        try:
            return self[k]
        except KeyError:
            return d

    def get_color(self, k, d=None):
        color = self.get(k, d)

        return get_hex(color)

    def convert_border(self):
        converter_dict = {
            'border': 'border-top-width: {0}px; border-top-style: {1}; border-top-color: {2}; '
                      'border-bottom-width: {0}px; border-bottom-style: {1}; border-bottom-color: {2}; '
                      'border-left-width: {0}px; border-left-style: {1}; border-left-color: {2}; '
                      'border-right-width: {0}px; border-right-style: {1}; border-right-color: {2}; ',
            'border-width': 'border-top-width: {0}px; '
                      'border-bottom-width: {0}px; '
                      'border-left-width: {0}px; '
                      'border-right-width: {0}px; ',
            'border-style': 'border-top-style: {0}; '
                      'border-bottom-style: {0}; '
                      'border-left-style: {0}; '
                      'border-right-style: {0} ;',
            'border-color': 'border-top-color: {0}; '
                      'border-bottom-color: {0}; '
                      'border-left-color: {0}; '
                      'border-right-color: {0}; ',
            'border-top': 'border-top-width: {0}px; '
                          'border-top-style: {1}; '
                          'border-top-color: {2}; ',
            'border-bottom': 'border-bottom-width: {0}px; '
                             'border-bottom-style: {1}; '
                             'border-bottom-color: {2}; ',
            'border-left': 'border-left-width: {0}px; '
                           'border-left-style: {1}; '
                           'border-left-color: {2}; ',
            'border-right': 'border-right-width: {0}px; '
                            'border-right-style: {1}; '
                            'border-right-color: {2}; ',
        }
        styles_to_update = ''
        for key in self.keys():
            new_value = converter_dict.get(key)
            if new_value:
                value_list = self[key].split()
                try:
                    first_arg = extract_first_int_from_str(value_list[0])  # width
                except IntNotFoundException:
                    first_arg = value_list[0]
                second_arg = value_list[1] if len(value_list) > 1 else 'solid'  # style
                third_arg = value_list[2] if len(value_list) > 2 else '#000000'  # color

                new_value = new_value.format(first_arg, second_arg, third_arg)
                styles_to_update += new_value
        if styles_to_update:
            new_value_dict = style_string_to_dict(styles_to_update)
            self.update(new_value_dict)

    def convert_style(self):
        converter_dict = {
            'background': 'background-color: {0}'
        }

        styles_to_update = ''
        for key in self.keys():
            new_value = converter_dict.get(key)
            if new_value:
                value_list = self[key].split()
                first_arg = value_list[0]
                new_value = new_value.format(first_arg)
                styles_to_update += new_value
        if styles_to_update:
            new_value_dict = style_string_to_dict(styles_to_update)
            self.update(new_value_dict)


class Element(object):
    def __init__(self, element, parent=None):
        self.element = element

        if self.__class__ != TableBody or self.element.tag != 'table':
            self._attribs_to_style_attrib()

        self.number_format = None
        parent_style = parent.style_dict if parent else None
        self.style_dict = StyleDict(style_string_to_dict(element.get('style', '')), parent=parent_style)
        self._style_cache = None

    def _attribs_to_style_attrib(self):
        new_styles = ''
        for attr_name, attr_value in self.element.attrib.items():
            new_attr_name = STYLES_CONVERTER_DICT.get(attr_name)
            if new_attr_name:
                new_styles += f'{new_attr_name}: {attr_value};'

        if not new_styles:
            return

        if not self.element.attrib.get('style'):
            self.element.attrib['style'] = new_styles
        else:
            self.element.attrib['style'] = new_styles + self.element.attrib['style']

    def style(self):
        if not self._style_cache:
            self._style_cache = style_dict_to_named_style(
                self.style_dict,
                number_format=self.number_format
            )
        return self._style_cache

    def get_dimension(self, dimension_key):
        dimension = self.style_dict.get(dimension_key)
        if dimension:
            if dimension[-2:] in ['px', 'em', 'pt', 'in', 'cm']:
                dimension = dimension[:-2]

            dimension = float(dimension)
        return dimension


class Table(Element):
    def __init__(self, table):
        super(Table, self).__init__(table)
        table_head = table.find('thead')
        self.head = TableHead(table_head, parent=self) if table_head is not None else None
        table_body = table.find('tbody')
        self.body = TableBody(table_body if table_body is not None else table, parent=self)

    def _attribs_to_style_attrib(self):
        new_styles = ''

        own_styles_converter_dict = {
            'bgcolor': 'background-color: {0};',
            'cellpadding': 'padding: {0}px;',
            'border': 'border-top-width: {0}px; border-top-style: solid; border-top-color: #000000;'
                      'border-bottom-width: {0}px; border-bottom-style: solid; border-bottom-color: #000000;'
                      'border-left-width: {0}px; border-left-style: solid; border-left-color: #000000;'
                      'border-right-width: {0}px; border-right-style: solid; border-right-color: #000000;'
        }

        for attr_name, attr_value in self.element.attrib.items():
            style_name = own_styles_converter_dict.get(attr_name)
            if style_name:
                new_styles += style_name.format(
                    attr_value
                    if attr_name != 'border' else extract_first_int_from_str(attr_value)
                )

        if not new_styles:
            return

        if not self.element.attrib.get('style'):
            self.element.attrib['style'] = new_styles
        else:
            self.element.attrib['style'] = new_styles + self.element.attrib['style']


class TableHead(Element):
    def __init__(self, head, parent=None):
        super(TableHead, self).__init__(head, parent=parent)
        self.rows = [TableRow(tr, parent=self) for tr in head.findall('tr')]
        self.cell_padding = self.get_dimension('padding') or 0


class TableBody(Element):
    def __init__(self, body, parent=None):
        super(TableBody, self).__init__(body, parent=parent)
        self.rows = [TableRow(tr, parent=self) for tr in body.findall('tr')]
        self.cell_padding = self.get_dimension('padding') or 0


class TableRow(Element):
    def __init__(self, tr, parent=None):
        super(TableRow, self).__init__(tr, parent=parent)
        self.cells = [TableCell(td, parent=self) for td in tr.findall('th') + tr.findall('td')]


class TableCell(Element):
    CELL_TYPES = {'TYPE_STRING', 'TYPE_FORMULA', 'TYPE_NUMERIC', 'TYPE_BOOL', 'TYPE_CURRENCY', 'TYPE_PERCENTAGE',
                  'TYPE_NULL', 'TYPE_INLINE', 'TYPE_ERROR', 'TYPE_FORMULA_CACHE_STRING', 'TYPE_INTEGER'}

    def __init__(self, cell, parent=None):
        self.cell = cell
        print(cell.text)
        print(cell.tail)
        self.value = self.element_to_string()
        super(TableCell, self).__init__(self.cell, parent=parent)
        self.number_format = self.get_number_format()
        self.style_dict.convert_border()
        self.style_dict.convert_style()

    def data_type(self):
        cell_types = self.CELL_TYPES & set(self.element.get('class', '').split())
        if cell_types:
            if 'TYPE_FORMULA' in cell_types:
                # Make sure TYPE_FORMULA takes precedence over the other classes in the set.
                cell_type = 'TYPE_FORMULA'
            elif cell_types & {'TYPE_CURRENCY', 'TYPE_INTEGER', 'TYPE_PERCENTAGE'}:
                cell_type = 'TYPE_NUMERIC'
            else:
                cell_type = cell_types.pop()
        else:
            cell_type = 'TYPE_STRING'
        return getattr(openpyxl_cell, cell_type)

    def get_number_format(self):
        if 'TYPE_CURRENCY' in self.element.get('class', '').split():
            return FORMAT_CURRENCY_USD_SIMPLE
        if 'TYPE_INTEGER' in self.element.get('class', '').split():
            return '#,##0'
        if 'TYPE_PERCENTAGE' in self.element.get('class', '').split():
            return FORMAT_PERCENTAGE
        if 'TYPE_DATE' in self.element.get('class', '').split():
            return FORMAT_DATE_MM_DD_YYYY
        if self.data_type() == openpyxl_cell.TYPE_NUMERIC:
            try:
                int(self.value)
            except ValueError:
                return '#,##0.##'
            else:
                return '#,##0'

    def format(self, cell):
        cell.style = self.style()
        data_type = self.data_type()
        if data_type:
            try:
                cell.data_type = data_type
            except AttributeError:
                pass

    def element_to_string(self):
        return self._element_to_string(self.cell)

    @staticmethod
    def extract_styles_from_font(font_tag):
        style_dict = StyleDict(style_string_to_dict(font_tag.get('style', '')))

        color = style_dict.get('color')
        size = style_dict.get('font-size')

        color = font_tag.attrib.get('color') if not color else color
        size = font_tag.attrib.get('size') if not size else size

        color = get_hex(color)
        try:
            size = extract_first_int_from_str(size)
        except IntNotFoundException:
            pass

        return color, size

    def _element_to_string(self, el):
        text_blocks = CellRichText()
        for x in el.iterchildren():
            child_text_blocks = self._element_to_string(x)
            text_blocks += child_text_blocks

        text = el.text if el.text else ''
        tail = el.tail if el.tail else ''

        text = re.sub(" +", " ", text)
        tail = re.sub(" +", " ", tail)

        font_style = InlineFont()
        if el.tag == 'font':
            color, size = self.extract_styles_from_font(el)
            font_style = InlineFont(color=color, sz=size)
        elif el.tag == 'b':
            font_style = InlineFont(b=True)

        result = CellRichText()

        result.append(TextBlock(font=font_style, text=text))
        for text_block in text_blocks:
            if font_style.b and not text_block.font.b:
                text_block.font.b = font_style.b if font_style.b and not text_block.font.b else text_block.font.b
            elif font_style.color and not text_block.font.color:
                text_block.font.color = font_style.color
        result += text_blocks

        if tail:
            result.append(TextBlock(font=InlineFont(), text=tail))

        return result
